from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium import webdriver
import pytest
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import openpyxl



@pytest.mark.usefixtures("setup")
class Locators:
    def __init__(self, driver):
        self.driver = driver

        self.radio_button_xpath = "//div[@id = 'radio-btn-example']/fieldset/label[2]/input"
        self.seggestion1_xpath = "//*[@id='autocomplete']"
        self.dro_down_optio = "//*[@id='dropdown-class-example']"
        self.check_box = "//*[@id='checkBoxOption1']"
        self.windo_switch = "//*[@id='openwindow']"
        self.choice = ["us", "eth", "au", "eng", "eri"]
        self.alert_xpath = "//*[@id='alertbtn']"
        self.table_amount = "//div[@class = 'tableFixHead']/table/tbody/tr/td[4]"
        self.mouse_hover = "//*[@id='mousehover']"
        self.new_tab = "//*[@id='opentab']"
        self.iframe_xpath = "//*[@id='courses-iframe']"

        self.book = openpyxl.load_workbook("home_page.xlsx")
        self.home_sheet = self.book.active


    #
    # # @pytest.mark.parametrize("num, result", [(1, 3), (2, 5)])
    # @pytest.mark.parametrize("num1,num2", [("ma", "ja"),
    #                                        ("ka", "fa"),
    #                                        ("Admin", "admin123")])
    def test_autoseggest(self):
        self.driver.find_element(By.XPATH, self.seggestion1_xpath).send_keys("ga")
        time.sleep(2)
        j = self.driver.find_element(By.XPATH, self.seggestion1_xpath).get_attribute("value")
        print("User entered", j, "on the country field")
        self.home_sheet.cell(row = 1, column=1, value="test_autoseggest")
        self.home_sheet.cell(row = 1, column=2, value=j)

        self.book.save("home_page.xlsx")

    def test_radio_button(self):
        self.driver.find_element(By.XPATH, self.radio_button_xpath).click()
        isselected = self.driver.find_element(By.XPATH, self.radio_button_xpath).is_selected()
        self.home_sheet.cell(row=2, column=1, value="test_radio_button selected1")
        self.home_sheet.cell(row=2, column=2, value=isselected)

        self.book.save("home_page.xlsx")

    def test_dro_dow(self):
        sel = self.driver.find_element(By.XPATH, self.dro_down_optio)
        select = Select(sel)
        select.select_by_visible_text("Option1")
        selected_option_dropdown = sel.get_attribute("value")

        self.home_sheet.cell(row=3, column=1, value="selected_option_dropdown")
        self.home_sheet.cell(row=3, column=2, value=selected_option_dropdown)

        self.book.save("home_page.xlsx")


    def test_check_box(self):
        self.driver.find_element(By.XPATH, self.check_box).click()

    def test_win(self):
        current_handle = self.driver.current_window_handle
        print("base_window", current_handle)
        self.driver.find_element(By.XPATH, self.windo_switch).click()
        time.sleep(2)
        all_handles = self.driver.window_handles

        for ac in all_handles:
            print(ac)
            if ac != current_handle:
                self.driver.switch_to.window(ac)
                self.home_sheet.cell(row=5, column=1, value="New driver window")
                self.home_sheet.cell(row=5, column=2, value=self.driver.current_window_handle)
                self.home_sheet.cell(row=5, column=3, value=self.driver.title)

                self.book.save("home_page.xlsx")


                time.sleep(5)
                self.driver.find_element(By.XPATH, "//*[@id='homepage']/div[4]/div[2]/div/div/div/span/div/div[6]/div/div/button").click()
                time.sleep(3)
                self.driver.find_element(By.XPATH, "//*[@id='homepage']/header/div[2]/div/nav/ul/li[2]/a").click()
                print(self.driver.title)

                self.driver.close()

                self.driver.switch_to.window(current_handle)

        self.home_sheet.cell(row=4, column=1, value="Current driver window")
        self.home_sheet.cell(row=4, column=2, value=current_handle)

        self.book.save("home_page.xlsx")

    def test_nw_windo(self, text):
        self.driver.find_element(By.XPATH, "//*[@id='name']").send_keys(text)




    def test_alert(self):
        self.driver.find_element(By.XPATH, self.alert_xpath).click()
        alert_obl = self.driver.switch_to.alert
        time.sleep(3)
        alert_obl.accept()
        time.sleep(3)

    def test_table_amount(self):
        total = 0
        tamount = self.driver.find_elements(By.XPATH, self.table_amount)
        for ac in tamount:
            print(ac.text)
            total += int(ac.text)
            print("the total of all  those numbers is", total)
            # assert 296 == total, "They dont matcj"

    def test_mouse_hover(self):
        mouse_h = self.driver.find_element(By.XPATH, self.mouse_hover)
        mouse_h.click()
        mouse_act = ActionChains(self.driver)
        self.driver.execute_script("window.scrollBy(0,500)", "")
        # self.driver.execute_script("arguments[0].scrollIntoView();", mouse_h)
        mouse_act.move_to_element(mouse_h)

        time.sleep(5)

        mouse_act.move_to_element(mouse_h)
    # def test_new_tab(self):
    #
    #     current_w = self.driver.current_window_handle
    #     print("Base window", current_w)
    #
    #     self.driver.find_element(By.XPATH, self.new_tab).click()
    #     all_tabs = self.driver.window_handles
    #     for ac in all_tabs:
    #         # print(ac)
    #         if ac != current_w:
    #             print(ac)
    #             self.driver.switch_to.window(ac)
    #
    #             self.driver.close()
    #
    #     print("Running it on jenkins")
    #     time.sleep(2)

    def test_ifrm(self):
        iframe = self.driver.find_element(By.XPATH, self.iframe_xpath)
        self.driver.switch_to.frame(iframe)

    def test_practice(self):
        self.driver.find_element(By.XPATH, "/html/body/div/header/div[3]/div/div/div[2]/nav/div[2]/ul/li[7]/a").click()




















